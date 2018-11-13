#!/usr/bin/env python
# -*- coding: utf-8 -*-

import paramiko
import time
import wmi
import re
from multiprocessing import Process
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side


class Myprocess(Process):
    def __init__(self, app, hostname, username, password, txt_fname, os_type):
        # 定义构造函数
        super(Myprocess, self).__init__()
        # Process.__init__(self)
        self.app = app
        self.hostname = hostname
        self.username = username
        self.password = password
        self.txt_fname = txt_fname
        self.os_type = os_type

    def execute_cmd(self, app, hostname, username, password, txt_fname, os_type):
        #########################################
        #    远程到服务器上执行命令获取状态值存入文件
        ########################################
        cmd_res = []
        cmd_shell = []  # 针对不同主机IP获取不同的shell命令
        with open('./cmd/' + os_type + '_cmd', 'r', encoding='utf-8') as f:
            for i in f.readlines():
                if not i.startswith('#'):
                    cmd_shell.append(i.strip('\n'))
        print(hostname)
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy)
        try:
            ssh.connect(hostname=hostname, username=username, password=password)
            for x in cmd_shell:
                stdin, stdout, stderr = ssh.exec_command(x)
                res = stdout.read().decode().strip('\n')
                cmd_res.append(res)
            ssh.close()

            with open(txt_fname, 'a', encoding='utf-8') as file:
                file.write(app + '  ' + hostname + '  ' + str(cmd_res) + '\n')
        except Exception as e:
            with open(txt_fname, 'a', encoding='utf-8') as file:
                file.write(app + '  ' + hostname + '  ' + str(e) + '\n')

    def run(self):
        self.execute_cmd(self.app, self.hostname, self.username, self.password, self.txt_fname, self.os_type)


def WriteEecel(txt_fname, xls_fname):
    #################################
    #    导入列表数据至csv文件
    #################################

    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'server check'  # 设置worksheet的标题

    # 设置列宽
    ws1.column_dimensions['A'].width = 13
    ws1.column_dimensions['B'].width = 17
    ws1.column_dimensions['C'].width = 10.89
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 8.78
    ws1.column_dimensions['F'].width = 13.11
    ws1.column_dimensions['G'].width = 5

    # 设置表头
    ws1['A1'] = '应用程序'
    ws1['B1'] = 'IP地址'
    ws1['C1'] = 'CPU使用率%'
    ws1['D1'] = '内存使用率%'
    ws1['E1'] = '/使用率%'
    ws1['F1'] = '/home使用率%'
    ws1['G1'] = '备注'

    # 设置填充样式
    fill_gray = PatternFill("solid", fgColor="CDCDC1")
    fill_red = PatternFill("solid", fgColor="FF0000")

    # 设置边框
    left, right, top, bottom = [Side(style='thin', color='000000')] * 4
    border = Border(left=left, right=right, top=top, bottom=bottom)

    with open(txt_fname, 'r', encoding='utf-8') as f:
        csv_ip_list = []
        row = 2
        for line in f.readlines():
            datas = line.split('  ')
            csv_app = datas[0]
            csv_ip = datas[1]
            if re.search('Error', datas[2]):
                # 如果远程未连上，数据为空
                csv_body = [csv_app, csv_ip, '', '', '', '']
            else:
                csv_data = eval(datas[2])
                csv_body = [csv_app, csv_ip, csv_data[0], csv_data[1], csv_data[2], csv_data[3]]
            csv_ip_list.append(csv_ip)
            # 导入数据
            for col, i in zip(range(1, 7), range(0, len(csv_body))):
                ws1.cell(row, col, csv_body[i])
                if i == 2 and csv_body[i] != '':
                    if float(csv_body[i]) > 90:  # cpu超过90%标红
                        ws1.cell(row, col).fill = fill_red

                if i == 3 and csv_body[i] != '':  # 内存超过90%标红
                    if float(csv_body[i]) > 90:
                        ws1.cell(row, col).fill = fill_red

                if i == 4 and csv_body[i] != '':  # 根分区超过80%标红
                    if float(csv_body[i]) > 80:
                        ws1.cell(row, col).fill = fill_red

                if i == 5 and csv_body[i] != '':  # 家目录超过80%标红
                    if float(csv_body[i]) > 80:
                        ws1.cell(row, col).fill = fill_red
            row += 1

    # 表头调用填充样式
    for col in 'ABCDEFG':
        ws1[col + '1'].fill = fill_gray

    # 调用边框
    for row in range(1, len(csv_ip_list) + 2):
        for col in 'ABCDEFG':
            ws1[col + str(row)].border = border

    wb.save(xls_fname)


def get_win(ip, user, password, app):
    try:
        c = wmi.WMI(computer=ip, user=user, password=password)

        # cpu 使用率
        cpu_usage = str(c.Win32_Processor()[0].LoadPercentage)

        # 内存使用率
        TotalMemory = int(c.Win32_ComputerSystem()[0].TotalPhysicalMemory) / 1024
        FreeMemory = int(c.Win32_OperatingSystem()[0].FreePhysicalMemory)
        memory_usage = str(int((TotalMemory - FreeMemory) / TotalMemory * 100))

        # 硬盘使用率bit
        TotalDisk = int(int(c.Win32_LogicalDisk(DeviceID="C:")[0].Size) / 1024 / 1024 / 1024)
        FreeDisk = int(int(c.Win32_LogicalDisk(DeviceID="C:")[0].FreeSpace) / 1024 / 1024 / 1024)
        UseDisk_per = str(int(100 - (FreeDisk) / TotalDisk * 100))

        # 监控数据列表
        win_res = [cpu_usage, memory_usage, UseDisk_per, '']

        with open(txt_fname, 'a', encoding='utf-8') as file:
            file.write(app + '  ' + ip + '  ' + str(win_res) + '\n')
    except Exception as e:

        with open(txt_fname, 'a', encoding='utf-8') as file:
            file.write(app + '  ' + ip + '  ' + str(e) + '\n')


if __name__ == '__main__':

    start = time.time()

    #  以时间戳定义txt日志以及xls日志文件名
    time_local = time.strftime("%Y%m%d_%H%M%S", time.localtime())
    txt_fname = './cmd_res/' + time_local + '.txt'
    xls_fname = './logs/' + time_local + '.xlsx'

    #  多进程获取主机巡检值
    with open('hosts.info', 'r', encoding='utf-8') as f:
        processes = []
        for line in f.readlines():
            line_list = line.split()
            app = line_list[0]
            hostname = line_list[1]
            username = line_list[2]
            password = line_list[3]
            os_type = line_list[4]
            if "centos" in os_type:
                p = Myprocess(app, hostname, username, password, txt_fname,os_type)
                processes.append(p)
                p.start()

                for p in processes:
                    # 等待进程退出
                    p.join()
            elif os_type == "windows":
                get_win(ip=hostname, user=username, password=password, app=app)

    WriteEecel(txt_fname, xls_fname)

    end = time.time()
    print('Task  runs %0.2f seconds.' % (end - start))
