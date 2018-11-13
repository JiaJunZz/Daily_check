#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/10/24 11:12
# @Author  : ZJJ
# @Email   : 597105373@qq.com

import paramiko
import xlrd
import time
import wmi
import shutil
import re
from xlutils3.copy import copy
from multiprocessing import Process
from openpyxl import Workbook,styles,formatting
from openpyxl.styles import  PatternFill,Border,Side


time_local = time.strftime("%Y%m%d_%H%M%S", time.localtime())
xls_fname = time_local + '.xlsx'

wb = Workbook()
ws1 = wb.active
ws1.title = 'server check'  # 设置worksheet的标题

# 设置表头
ws1['A1'] = '应用程序'
ws1['B1'] = 'IP地址'
ws1['C1'] = '进程名'
ws1['D1'] = 'CPU使用率%'
ws1['E1'] = '内存使用率%'
ws1['F1'] = '/空闲率%'
ws1['G1'] = '/home空闲率%'
ws1['H1'] = '备注'

# 设置填充样式
fill_gray = PatternFill("solid", fgColor="CDCDC1")
fill_red = PatternFill("solid", fgColor="FF0000")

# 设置边框
left, right, top, bottom = [Side(style='thin',color='000000')]*4
border = Border(left=left, right=right, top=top, bottom=bottom)

with open('20181024_192258.txt','r',encoding='utf-8') as f:
    csv_ip_list = []
    row = 2
    for line in f.readlines():
        datas = line.split('  ')
        csv_app = datas[0]
        csv_ip = datas[1]
        csv_process = datas[2]
        if re.search('Error',datas[3]):
            # 如果远程未连上，数据为空
            csv_body = [csv_app, csv_ip, csv_process, '', '', '', '']
        else:
            csv_data = eval(datas[3])
            csv_body = [csv_app, csv_ip, csv_process, csv_data[0], csv_data[1], csv_data[2], csv_data[3]]
        csv_ip_list.append(csv_ip)
        # 导入数据
        for col,i in zip(range(1, 8),range(0, len(csv_body))):
            ws1.cell(row, col, csv_body[i])
            if i == 3  and csv_body[i] != '':
                if float(csv_body[i]) > 90:   # cpu超过90%标红
                    ws1.cell(row, col).fill = fill_red

            if i == 4 and csv_body[i] != '':  # 内存超过90%标红
                if float(csv_body[i]) > 90:
                    ws1.cell(row, col).fill = fill_red

            if i == 5 and csv_body[i] != '':  # 根分区超过80%标红
                if float(csv_body[i]) > 80:
                    ws1.cell(row, col).fill = fill_red

            if i == 6 and csv_body[i] != '':  # 家目录超过80%标红
                if float(csv_body[i]) > 80:
                    ws1.cell(row, col).fill = fill_red
        row += 1

# 表头调用填充样式
for col in 'ABCDEFGH':
    ws1[col+'1'].fill = fill_gray

# 调用边框
for row in range(1, len(csv_ip_list)+2):
    for col in 'ABCDEFGH':
        ws1[col + str(row)].border = border

wb.save(xls_fname)